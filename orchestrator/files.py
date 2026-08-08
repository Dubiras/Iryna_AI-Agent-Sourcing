# © 2026 Iryna Subbotina. All Rights Reserved. Proprietary — not open source. See LICENSE/NOTICE. Unauthorized copying, redistribution, or claim of authorship is prohibited.
"""File content extraction for PDF, XLSX, DOC/DOCX, and audio."""

import io
import os
from pathlib import Path


SUPPORTED_EXTENSIONS = {
    ".pdf", ".xlsx", ".xls", ".doc", ".docx",
    ".mp3", ".m4a", ".ogg", ".wav", ".oga", ".flac",
}


def is_supported(filename: str) -> bool:
    return Path(filename).suffix.lower() in SUPPORTED_EXTENSIONS


def is_audio_file(filename: str) -> bool:
    return Path(filename).suffix.lower() in {".mp3", ".m4a", ".ogg", ".wav", ".oga", ".flac"}


async def extract(file_bytes: bytes, filename: str) -> str:
    """Extract text from a file. Returns extracted text or raises ValueError."""
    ext = Path(filename).suffix.lower()

    if ext == ".pdf":
        return _extract_pdf(file_bytes, filename)
    elif ext in {".xlsx", ".xls"}:
        return _extract_xlsx(file_bytes, filename)
    elif ext == ".docx":
        return _extract_docx(file_bytes, filename)
    elif ext == ".doc":
        return _extract_doc(file_bytes, filename)
    elif ext in {".mp3", ".m4a", ".ogg", ".wav", ".oga", ".flac"}:
        return await _transcribe_audio(file_bytes, filename)
    else:
        raise ValueError(f"Непідтримуваний формат: {ext}")


def _extract_pdf(data: bytes, filename: str) -> str:
    try:
        import pypdf
        reader = pypdf.PdfReader(io.BytesIO(data))
        pages = [page.extract_text() or "" for page in reader.pages]
        text = "\n\n".join(p for p in pages if p.strip())
        if not text.strip():
            return f"[PDF '{filename}' не містить тексту або захищений]"
        return f"[PDF: {filename}]\n\n{text}"
    except Exception as e:
        raise ValueError(f"Помилка читання PDF: {e}")


def _extract_xlsx(data: bytes, filename: str) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    rows.append("\t".join(cells))
            if rows:
                parts.append(f"--- Аркуш: {sheet} ---\n" + "\n".join(rows))
        wb.close()
        if not parts:
            return f"[XLSX '{filename}' порожній]"
        return f"[Excel: {filename}]\n\n" + "\n\n".join(parts)
    except Exception as e:
        raise ValueError(f"Помилка читання XLSX: {e}")


def _extract_docx(data: bytes, filename: str) -> str:
    try:
        import docx
        doc = docx.Document(io.BytesIO(data))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        if not paragraphs:
            return f"[DOCX '{filename}' порожній]"
        return f"[Word: {filename}]\n\n" + "\n".join(paragraphs)
    except Exception as e:
        raise ValueError(f"Помилка читання DOCX: {e}")


def _extract_doc(data: bytes, filename: str) -> str:
    # Old .doc format — try antiword via subprocess or fallback message
    try:
        import subprocess
        result = subprocess.run(
            ["antiword", "-"],
            input=data, capture_output=True, timeout=15
        )
        if result.returncode == 0 and result.stdout:
            text = result.stdout.decode("utf-8", errors="replace")
            return f"[Word: {filename}]\n\n{text}"
    except Exception:
        pass
    return (
        f"[Файл '{filename}' у старому форматі .doc]\n"
        "Збережи як .docx і надішли знову — тоді зможу прочитати."
    )


async def _transcribe_audio(data: bytes, filename: str) -> str:
    from voice import transcribe
    transcript = await transcribe(data, filename=filename)
    if not transcript:
        raise ValueError("Не вдалося транскрибувати аудіо")
    return f"[Аудіо: {filename}]\n\n{transcript}"
